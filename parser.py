import os
from ..imc import models
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

IMC_APPLICATION_MAPPING = {
    "Part": "part_number",
    "Make": "make_id",
    "Note": "note",
    "MfrLabel": "mfr_label",
    "Positions": "position_id",
    "EngineBase": "engine_base_id",
    "TransmissionControlType": "transmission_control_type_id",
    "EngineDesignation": "engine_designation_id",
    "BodyType": "body_type_id",
    "BaseVehicle": "base_vehicle_id",
    "SteeringSystem": "steering_system_id",
    "VehicleType": "vehicle_type_id",
    "Model": "model_id",
    "AssetName": "asset_name_id",
    "PartType": "part_type_id",
    "Qty": "qty",
    "FuelType": "fuel_type_id",
}

IMC_ASSET_MAPPING = {
    "Note": "note",
    "Make": "make",
    "Model": "model",
    "SubModel": "submodel",
    "EngineBase": "engine_base",
    "AssetName": "asset_name"
}

IMC_DIGITAL_ASSET_MAPPING = {
    "FileName": "file_name",
    "AssetDetailType": "asset_detail_type",
    "FileType": "file_type",
    "Representation": "representation",
    "FileSize": "file_size",
    "Resolution": "resolution",
    "ColorMode": "color_mode",
    "Background": "background",
    "OrientationView": "orientation_view",
    "AssetDescription": "asset_description",
    "FilePath": "file_path",
    "URI": "uri",
    "FileDateModified": "file_date_modified",
    "EffectiveDate": "effective_date",
    "ExpirationDate": "expiration_date",
    "Country": "country"
}
TAG_MAPPING = {
    'App': IMC_APPLICATION_MAPPING,
    'Asset': IMC_ASSET_MAPPING,
    'DigitalAsset': IMC_DIGITAL_ASSET_MAPPING,
}
TAG_MODEL = {
    'App': models.IMCApplications,
    'Asset': models.Asset,
    'DigitalAsset': models.DigitalAsset,
}
MODELS_WITH_ID = [models.DigitalAsset, models.IMCPrice, models.Header]

IMC_PRICE_MAPPING = {
    u'IMC UNSPACED': 'imc',
    u'CORE': 'core',
    u'DESCRIPTION': 'description',
    u'WEIGHT': 'weight',
    u'BRAND': 'brand',
    u'DEALER LIST': 'dealer_list',
    u'CLASS CODE': 'class_code',
    u'NET': 'net',
    u'JPN UNSPACED': 'jpn',
    u'UOM': 'uom',
    u'MANUFACTURER': 'mfr'}


def xml_to_dict(element):
    """
    Convert simple xml to dict
    """
    dict_data = {}
    for child in element:
        attr = []
        if child.attrib == {}:
            dict_data.update({child.tag: child.text})
        else:
            for attrib_name, attrib_value in child.attrib.items():
                attr.append({attrib_name: attrib_value})
            dict_data.update({
                child.tag: {'attribute': attr, 'value': child.text}
            })
    return dict_data


class IMCSaver(obj):

    def get_mapping(self, root_tag):
        """
        Return field mapping depends on xml tag
        """
        return TAG_MAPPING[root_tag.tag]

    def create_mapping(self, ModelClass):
        """
        Create field mapping for ModelClass
        """
        mapping = []
        for field in ModelClass._meta.fields:
            mapping.append(field.get_attname_column())
        if ModelClass in MODELS_WITH_ID:
            return mapping[1:]
        return mapping

    def save_header(self, root, ModelClass, tag_name):
        """
        Save simple header
        """
        obj = ModelClass()
        child = root.find(tag_name)
        model_mapping = self.create_mapping(ModelClass)
        for field_row in model_mapping:
            tag = child.find(field_row[1])
            if tag is None:
                continue
            setattr(obj, field_row[0], tag.text)
        return obj.save()

    def save_data(self, root, ModelClass, tag_name, header):
        """
        Create ModelClass objects
        """
        data = None
        objects = []
        for child in root.findall(tag_name):
            obj = ModelClass()
            for app_tag in child:
                tag_to_field_mapping = self.get_mapping(child)
                if app_tag.tag in tag_to_field_mapping:
                    if app_tag.attrib != {}:
                        setattr(obj, tag_to_field_mapping[app_tag.tag], app_tag.attrib['id'])
                        continue
                    setattr(obj, tag_to_field_mapping[app_tag.tag], app_tag.text)
                else:
                    data = xml_to_dict(child)
            obj.id = child.attrib['id']
            obj.data = data
            obj.action = child.attrib['action']
            years = child.find('Years')
            if years:
                obj.year_from_id = years.attrib['from']
                obj.year_to_id = years.attrib['to']
            obj.header = header
            if tag_name == 'DigitalFileInformation':
                obj.language_code = child.attrib['LanguageCode']
                obj.asset_name = child.attrib['AssetName']
                dimension = tag_name.find("AssetDimensions")
                obj.asset_dimensions = xml_to_dict(dimension)
            objects.append(obj)

            if len(objects) > 500:
                ModelClass.objects.bulk_create(objects)
                objects = []

        if objects:
            ModelClass.objects.bulk_create(objects)

    def parse_and_save(self, path):
        """
        Parse file and save objects to database
        :param path: path to file Interamerican Motor Corporation_VW_Audi_2015-7-30_110124_FULL.XML
        :return:
        """
        tree = ET.parse(os.path.join(path, 'Interamerican Motor Corporation_VW_Audi_2015-7-30_110124_FULL.XML'))
        root = tree.getroot()
        header = self.save_header(root, models.Header, 'Header')
        for tag_name, ModelClass in TAG_MODEL.items():
            self.save_data(root, ModelClass, tag_name, header)
            if tag_name == "DigitalAsset":
                for child in root:
                    self.save_data(child, ModelClass, tag_name, header)

    def get_price(self, path):
        """
        Read information about price from xlsx
        """
        wb = load_workbook(filename=path, read_only=True)
        worksheet = wb.worksheets[0].iter_rows()
        row_names = {}
        for row in worksheet:
            for cell in row:
                row_names[cell.column] = cell.value.strip()
            break

        for row in worksheet:
            entry = {}
            for cell in row:
                if cell.column not in row_names:
                    continue
                entry[row_names[cell.column]] = cell.value
            yield entry

    def create_imc_price(self, path):
        """
        Create IMCPrice objects from file
        """
        objects = []
        for row in self.get_price(path):
            obj = models.IMCPrice()
            for cell, val in row.items():
                if not cell in IMC_PRICE_MAPPING:
                    continue
                setattr(obj, IMC_PRICE_MAPPING[cell], val)
            objects.append(obj)
            if len(objects) >= 500:
                models.IMCPrice.objects.bulk_create(objects)
                objects = []
        if objects:
            models.IMCPrice.objects.bulk_create(objects)

